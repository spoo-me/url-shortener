from flask import Response, jsonify
from typing import Optional, Any, Dict
import io
import csv
import zipfile
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from dicttoxml import dicttoxml
from flask import send_file

from .stats import StatsQueryBuilder


class ExportBuilder:
    """Builder for exporting URL statistics in various formats (CSV, XLSX, JSON, XML)"""

    def __init__(self, owner_id, args: dict[str, Any]):
        self.owner_id = owner_id
        self.args = args
        self.error: Optional[tuple[Response, int]] = None

        # Export parameters
        self.format: Optional[str] = None
        self.stats_data: Optional[Dict[str, Any]] = None
        self.stats_builder: Optional[StatsQueryBuilder] = None

        # Allowed formats
        self.allowed_formats = {"csv", "xlsx", "json", "xml"}

    def _fail(self, body: dict, status: int) -> "ExportBuilder":
        """Set error state"""
        self.error = (jsonify(body), status)
        return self

    def parse_format(self) -> "ExportBuilder":
        """Parse and validate the export format parameter"""
        if self.error:
            return self

        format_raw = self.args.get("format", "").strip().lower()
        if not format_raw:
            return self._fail(
                {"error": "format parameter is required (csv, xlsx, json, xml)"}, 400
            )

        if format_raw not in self.allowed_formats:
            return self._fail(
                {
                    "error": f"invalid format - must be one of: {', '.join(self.allowed_formats)}"
                },
                400,
            )

        self.format = format_raw
        return self

    def parse_stats(self) -> "ExportBuilder":
        """
        Parse statistics using StatsQueryBuilder to avoid duplication.
        This reuses all the filtering, grouping, and validation logic.
        """
        if self.error:
            return self

        # Create and execute StatsQueryBuilder
        self.stats_builder = (
            StatsQueryBuilder(self.owner_id, self.args)
            .parse_auth_scope()
            .parse_scope_and_target()
            .parse_time_range()
            .parse_filters()
            .parse_group_by()
            .parse_metrics()
            .parse_timezone()
        )

        # Check if stats builder has any errors
        if self.stats_builder.error:
            self.error = self.stats_builder.error
            return self

        # Get the stats data by building the response
        try:
            response, status = self.stats_builder.build()
            if status != 200:
                self.error = (response, status)
                return self

            # Extract JSON data from response
            self.stats_data = response.get_json()
        except Exception as e:
            print(f"Error fetching stats for export: {e}")
            return self._fail({"error": "failed to fetch statistics"}, 500)

        return self

    def build_export(self) -> "ExportBuilder":
        """Prepare the export file (validation only, actual generation in send())"""
        if self.error:
            return self

        if not self.stats_data:
            return self._fail({"error": "no statistics data available"}, 500)

        return self

    def _export_to_csv(self) -> Response:
        """Export statistics to CSV format (zipped multiple files)"""
        output = io.BytesIO()

        with zipfile.ZipFile(
            output, mode="w", compression=zipfile.ZIP_DEFLATED
        ) as zipf:
            # Write summary CSV
            summary_data = self.stats_data.get("summary", {})
            time_range = self.stats_data.get("time_range", {})
            scope = self.stats_data.get("scope", "")
            timezone_val = self.stats_data.get("timezone", "UTC")

            with zipf.open("summary.csv", "w") as file:
                with io.TextIOWrapper(file, encoding="utf-8", newline="") as text_file:
                    writer = csv.writer(text_file)
                    writer.writerow(["Metric", "Value"])
                    writer.writerow(["Scope", scope])
                    writer.writerow(["Timezone", timezone_val])
                    writer.writerow(["Start Date", time_range.get("start_date", "N/A")])
                    writer.writerow(["End Date", time_range.get("end_date", "N/A")])
                    writer.writerow(
                        ["Total Clicks", summary_data.get("total_clicks", 0)]
                    )
                    writer.writerow(
                        ["Unique Clicks", summary_data.get("unique_clicks", 0)]
                    )
                    writer.writerow(
                        ["First Click", summary_data.get("first_click", "N/A")]
                    )
                    writer.writerow(
                        ["Last Click", summary_data.get("last_click", "N/A")]
                    )
                    writer.writerow(
                        [
                            "Avg Redirection Time (ms)",
                            summary_data.get("avg_redirection_time", 0),
                        ]
                    )

            # Write metrics CSVs
            metrics_data = self.stats_data.get("metrics", {})
            for metric_key, metric_values in metrics_data.items():
                # Parse metric_key like "clicks_by_time" to get dimension and metric
                parts = metric_key.split("_by_")
                if len(parts) == 2:
                    metric_name = parts[0]
                    dimension_name = parts[1]

                    filename = f"{metric_key}.csv"
                    with zipf.open(filename, "w") as file:
                        with io.TextIOWrapper(
                            file, encoding="utf-8", newline=""
                        ) as text_file:
                            writer = csv.writer(text_file)
                            # Header
                            writer.writerow([dimension_name.title(), metric_name])
                            # Data rows
                            for row in metric_values:
                                dim_value = row.get(dimension_name, "unknown")
                                metric_value = row.get(metric_name, 0)
                                writer.writerow([dim_value, metric_value])

        output.seek(0)
        return send_file(
            output,
            mimetype="application/zip",
            as_attachment=True,
            download_name="spoo-me-export.zip",
        )

    def _export_to_xlsx(self) -> Response:
        """Export statistics to Excel format (XLSX)"""
        output = io.BytesIO()
        wb = Workbook()

        # Bold font style
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center")

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_data = self.stats_data.get("summary", {})
        time_range = self.stats_data.get("time_range", {})
        scope = self.stats_data.get("scope", "")
        timezone_val = self.stats_data.get("timezone", "UTC")

        summary_rows = [
            ["Metric", "Value"],
            ["Scope", scope],
            ["Timezone", timezone_val],
            ["Start Date", time_range.get("start_date", "N/A")],
            ["End Date", time_range.get("end_date", "N/A")],
            ["Total Clicks", summary_data.get("total_clicks", 0)],
            ["Unique Clicks", summary_data.get("unique_clicks", 0)],
            ["First Click", summary_data.get("first_click", "N/A")],
            ["Last Click", summary_data.get("last_click", "N/A")],
            [
                "Avg Redirection Time (ms)",
                summary_data.get("avg_redirection_time", 0),
            ],
        ]

        for row in summary_rows:
            ws_summary.append(row)

        # Style summary sheet
        for cell in ws_summary["A"]:
            cell.font = bold_font
        for cell in ws_summary[1]:
            cell.font = bold_font
            cell.alignment = center_align
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 30

        # Metrics sheets
        metrics_data = self.stats_data.get("metrics", {})
        for metric_key, metric_values in metrics_data.items():
            # Parse metric_key like "clicks_by_time" to get dimension and metric
            parts = metric_key.split("_by_")
            if len(parts) == 2:
                metric_name = parts[0]
                dimension_name = parts[1]

                # Create safe sheet name (Excel has 31 char limit)
                sheet_name = f"{metric_name}_{dimension_name}"[:31]
                ws = wb.create_sheet(sheet_name)

                # Header
                ws.append([dimension_name.title(), metric_name])
                for cell in ws[1]:
                    cell.font = bold_font
                    cell.alignment = center_align

                # Data rows
                for row in metric_values:
                    dim_value = row.get(dimension_name, "unknown")
                    metric_value = row.get(metric_name, 0)
                    ws.append([dim_value, metric_value])

                # Set column widths
                ws.column_dimensions["A"].width = 25
                ws.column_dimensions["B"].width = 15

        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="spoo-me-export.xlsx",
        )

    def _export_to_json(self) -> Response:
        """Export statistics to JSON format"""
        output = io.StringIO()
        json.dump(self.stats_data, output, indent=2)
        output.seek(0)

        output_bytes = io.BytesIO(output.getvalue().encode())

        return send_file(
            output_bytes,
            mimetype="application/json",
            as_attachment=True,
            download_name="spoo-me-export.json",
        )

    def _export_to_xml(self) -> Response:
        """Export statistics to XML format"""
        # Convert dictionary to XML
        xml = dicttoxml(self.stats_data, custom_root="statistics", attr_type=False)

        # Create BytesIO object and write XML data to it
        output = io.BytesIO()
        output.write(xml)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/xml",
            as_attachment=True,
            download_name="spoo-me-export.xml",
        )

    def send(self) -> tuple[Response, int]:
        """Generate and send the export file"""
        if self.error:
            return self.error

        try:
            if self.format == "csv":
                return self._export_to_csv(), 200
            elif self.format == "xlsx":
                return self._export_to_xlsx(), 200
            elif self.format == "json":
                return self._export_to_json(), 200
            elif self.format == "xml":
                return self._export_to_xml(), 200
            else:
                return self._fail({"error": "unsupported format"}, 400)
        except Exception as e:
            print(f"Export generation error: {e}")
            return jsonify({"error": "failed to generate export"}), 500
