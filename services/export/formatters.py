"""
Export formatters — one class per serialisation format.

Each formatter owns its mimetype, filename, and serialisation logic.
Register new formats by creating a class implementing ExportFormatter
and adding it to default_formatters() — ExportService never changes.
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import datetime
from typing import Any

from dicttoxml import dicttoxml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.export.protocol import ExportFormatter


def _excel_safe(val: Any) -> Any:
    """Strip timezone info from datetimes for Excel compatibility."""
    if isinstance(val, datetime) and val.tzinfo is not None:
        return val.replace(tzinfo=None)
    return val


class JsonFormatter:
    mimetype = "application/json"
    filename = "spoo-me-export.json"

    def serialize(self, data: dict[str, Any]) -> bytes:
        return json.dumps(data, indent=4, default=str).encode("utf-8")


class XmlFormatter:
    mimetype = "application/xml"
    filename = "spoo-me-export.xml"

    def serialize(self, data: dict[str, Any]) -> bytes:
        return dicttoxml(data)


class CsvFormatter:
    """Produces a ZIP archive containing one CSV per metrics dimension."""

    mimetype = "application/zip"
    filename = "spoo-me-export-csv.zip"

    def serialize(self, data: dict[str, Any]) -> bytes:
        output = io.BytesIO()

        def _write_rows(
            zipf: zipfile.ZipFile,
            filename: str,
            rows: list[tuple[str, Any]],
        ) -> None:
            with (
                zipf.open(filename, "w") as f,
                io.TextIOWrapper(f, encoding="utf-8", newline="") as txt,
            ):
                writer = csv.writer(txt)
                for row in rows:
                    writer.writerow(row)

        with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            summary = data.get("summary", {})
            summary_rows: list[tuple[str, Any]] = [
                ("Field", "Value"),
                ("total_clicks", summary.get("total_clicks", 0)),
                ("unique_clicks", summary.get("unique_clicks", 0)),
                ("first_click", summary.get("first_click", "")),
                ("last_click", summary.get("last_click", "")),
                ("avg_redirection_time", summary.get("avg_redirection_time", 0)),
            ]
            _write_rows(zf, "summary.csv", summary_rows)

            for metric_key, metric_data in data.get("metrics", {}).items():
                if not isinstance(metric_data, list) or not metric_data:
                    continue
                headers = list(metric_data[0].keys())
                rows: list[tuple[str, Any]] = [tuple(headers)]  # type: ignore[list-item]
                for item in metric_data:
                    rows.append(tuple(item.get(h, "") for h in headers))  # type: ignore[misc]
                _write_rows(zf, f"{metric_key}.csv", rows)

        output.seek(0)
        return output.read()


class XlsxFormatter:
    """Produces an XLSX workbook with a Summary sheet and one sheet per metrics key."""

    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = "spoo-me-export.xlsx"

    def serialize(self, data: dict[str, Any]) -> bytes:
        wb = Workbook()
        bold = Font(bold=True)
        center = Alignment(horizontal="center")

        ws = wb.active
        ws.title = "Summary"
        summary = data.get("summary", {})
        ws.append(["Field", "Value"])
        for key, val in summary.items():
            ws.append([key, _excel_safe(val)])
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = center
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25

        for metric_key, metric_data in data.get("metrics", {}).items():
            if not isinstance(metric_data, list) or not metric_data:
                continue
            ws_dim = wb.create_sheet(title=metric_key[:31])
            headers = list(metric_data[0].keys())
            ws_dim.append(headers)
            for item in metric_data:
                ws_dim.append([_excel_safe(item.get(h, "")) for h in headers])
            for cell in ws_dim[1]:
                cell.font = bold
                cell.alignment = center
            ws_dim.column_dimensions["A"].width = 25

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


def default_formatters() -> dict[str, ExportFormatter]:
    """Standard formatter registry for the composition root."""
    return {
        "json": JsonFormatter(),
        "xml": XmlFormatter(),
        "csv": CsvFormatter(),
        "xlsx": XlsxFormatter(),
    }
