"""
Legacy stats and export routes.

GET  /stats                           → stats entry form (HTML)
POST /stats                           → redirect to /stats/<code> (HTML)
GET  /stats/<short_code>              → analytics page (HTML)
POST /stats/<short_code>              → analytics data (JSON)
GET  /export/<short_code>/<format>    → export page redirect (HTML) or file
POST /export/<short_code>/<format>    → export data (file download)
"""

from __future__ import annotations

import csv
import io
import json
import os
import zipfile
from datetime import datetime, timezone
from urllib.parse import unquote

from dicttoxml import dicttoxml
from fastapi import APIRouter, Depends, Request
from fastapi.responses import (
    JSONResponse,
    RedirectResponse,
    Response,
)
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from dependencies import get_db
from middleware.rate_limiter import limiter
from repositories.legacy.emoji_url_repository import EmojiUrlRepository
from repositories.legacy.legacy_url_repository import LegacyUrlRepository
from shared.datetime_utils import convert_to_gmt
from shared.logging import get_logger
from shared.validators import validate_emoji_alias
from shared.legacy_helpers import (
    add_missing_dates,
    calculate_click_averages,
    convert_country_data,
    get_stats_pipeline,
    top_four,
)

log = get_logger(__name__)

router = APIRouter(include_in_schema=False)

_TEMPLATE_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "templates"
)
templates = Jinja2Templates(directory=_TEMPLATE_DIR)


# ── Stats entry form ──────────────────────────────────────────────────────────


@router.api_route("/stats", methods=["GET", "POST"], include_in_schema=False)
@router.api_route("/stats/", methods=["GET", "POST"], include_in_schema=False)
@limiter.limit("20 per minute; 1000 per day")
async def stats_route(request: Request, db=Depends(get_db)) -> Response:
    """Stats entry form — GET renders the form; POST redirects to /stats/<code>."""
    host_url = str(request.base_url)

    if request.method == "GET":
        return templates.TemplateResponse(request, "stats.html", {"host_url": host_url})

    form = await request.form()
    short_code = form.get("short_code") or request.query_params.get("short_code")

    if not short_code:
        return JSONResponse(
            {"error": "Invalid Short Code, short code does not exist!"}, status_code=400
        )

    # Strip leading URL prefix (user may paste full short URL)
    short_code = short_code[short_code.rfind("/") + 1 :]
    password = form.get("password") or request.query_params.get("password")
    short_code = unquote(short_code)

    # Check existence and password gate
    if validate_emoji_alias(short_code):
        url_data = await EmojiUrlRepository(db["emojis"]).find_by_id(short_code)
    else:
        url_data = await LegacyUrlRepository(db["urls"]).find_by_id(short_code)

    if not url_data:
        log.info("legacy_stats_not_found", short_code=short_code)
        return templates.TemplateResponse(
            request,
            "stats.html",
            {
                "error": "Invalid Short Code, short code does not exist!",
                "url": short_code,
                "host_url": host_url,
            },
        )

    stored_password = url_data.password

    if not password and stored_password is not None:
        log.info("legacy_stats_password_required", short_code=short_code)
        return templates.TemplateResponse(
            request,
            "stats.html",
            {
                "password_error": (
                    f"{host_url}{short_code} is a password protected Url,"
                    " please enter the password to continue."
                ),
                "url": short_code,
                "host_url": host_url,
            },
        )

    if stored_password is not None and stored_password != password:
        log.warning("legacy_stats_password_incorrect", short_code=short_code)
        return templates.TemplateResponse(
            request,
            "stats.html",
            {
                "password_error": "Invalid Password! please enter the correct password to continue.",
                "url": short_code,
                "host_url": host_url,
            },
        )

    if stored_password is not None:
        return RedirectResponse(
            f"/stats/{short_code}?password={password}", status_code=302
        )
    return RedirectResponse(f"/stats/{short_code}", status_code=302)


# ── Analytics page / JSON ─────────────────────────────────────────────────────


@router.api_route(
    "/stats/{short_code}", methods=["GET", "POST"], include_in_schema=False
)
@limiter.limit("20 per minute; 1000 per day")
async def analytics(short_code: str, request: Request, db=Depends(get_db)) -> Response:
    """Analytics page — GET renders stats_view.html; POST returns JSON."""
    password = request.query_params.get("password") or (
        (await request.form()).get("password") if request.method == "POST" else None
    )
    short_code = unquote(short_code)
    host_url = str(request.base_url)
    pipeline = get_stats_pipeline(short_code)

    if validate_emoji_alias(short_code):
        url_data = await EmojiUrlRepository(db["emojis"]).aggregate(pipeline)
    else:
        url_data = await LegacyUrlRepository(db["urls"]).aggregate(pipeline)

    if not url_data:
        log.info(
            "legacy_analytics_not_found", short_code=short_code, method=request.method
        )
        if request.method == "GET":
            return templates.TemplateResponse(
                request,
                "error.html",
                {
                    "error_code": "404",
                    "error_message": "URL NOT FOUND",
                    "host_url": host_url,
                },
                status_code=404,
            )
        return JSONResponse(
            {"UrlError": "The requested Url never existed"}, status_code=404
        )

    if url_data.get("password") is not None:
        if password != url_data["password"]:
            log.warning("legacy_analytics_password_incorrect", short_code=short_code)
            if request.method == "POST":
                return JSONResponse(
                    {"PasswordError": "Invalid Password"}, status_code=400
                )
            if not password:
                return templates.TemplateResponse(
                    request,
                    "stats.html",
                    {
                        "url": short_code,
                        "password_error": (
                            f"{host_url}{short_code} is a password protected Url,"
                            " please enter the password to continue."
                        ),
                        "host_url": host_url,
                    },
                    status_code=400,
                )
            return templates.TemplateResponse(
                request,
                "stats.html",
                {
                    "url": short_code,
                    "password_error": "Invalid Password! please enter the correct password to continue.",
                    "host_url": host_url,
                },
                status_code=400,
            )

    # Expiry check
    if url_data.get("max-clicks") is not None:
        url_data["expired"] = url_data["total-clicks"] >= int(url_data["max-clicks"])
    else:
        url_data["expired"] = None

    if url_data.get("expiration-time") is not None:
        expiration_time = convert_to_gmt(url_data["expiration-time"])
        if expiration_time and expiration_time <= datetime.now(timezone.utc):
            url_data["expired"] = True

    url_data["short_code"] = short_code
    (
        url_data["average_daily_clicks"],
        url_data["average_weekly_clicks"],
        url_data["average_monthly_clicks"],
    ) = calculate_click_averages(url_data)
    url_data["average_redirection_time"] = url_data.get("average_redirection_time", 0)

    if url_data["counter"]:
        url_data = add_missing_dates("counter", url_data)
    if url_data.get("unique_counter"):
        url_data = add_missing_dates("unique_counter", url_data)

    if request.method == "POST":
        return JSONResponse(url_data)

    try:
        url_data["hyper_link"] = url_data["url"]
        url_data["sorted_country"] = convert_country_data(url_data["country"])
        url_data["sorted_referrer"] = json.dumps(top_four(url_data["referrer"]))
        url_data["sorted_os_name"] = top_four(url_data["os_name"])
        url_data["sorted_browser"] = top_four(url_data["browser"])
        url_data["sorted_unique_browser"] = top_four(url_data["unique_browser"])
        url_data["sorted_unique_os_name"] = top_four(url_data["unique_os_name"])
        url_data["sorted_unique_country"] = convert_country_data(
            url_data["unique_country"]
        )
        url_data["sorted_unique_referrer"] = json.dumps(
            top_four(url_data["unique_referrer"])
        )
        url_data["sorted_bots"] = top_four(url_data["bots"])
        url_data["analysis_data"] = {
            "average_daily_clicks": url_data["average_daily_clicks"],
            "average_weekly_clicks": url_data["average_weekly_clicks"],
            "average_monthly_clicks": url_data["average_monthly_clicks"],
        }
    except Exception:
        pass

    return templates.TemplateResponse(
        request, "stats_view.html", {"json_data": url_data, "host_url": host_url}
    )


# ── Export ────────────────────────────────────────────────────────────────────


@router.api_route(
    "/export/{short_code}/{fmt}", methods=["GET", "POST"], include_in_schema=False
)
@limiter.limit("10 per minute; 200 per day")
async def export(
    short_code: str, fmt: str, request: Request, db=Depends(get_db)
) -> Response:
    """Export URL analytics in JSON, CSV, XLSX, or XML format."""
    fmt = fmt.lower()
    password = request.query_params.get("password") or (
        (await request.form()).get("password") if request.method == "POST" else None
    )
    short_code = unquote(short_code)
    host_url = str(request.base_url)

    if fmt not in ("csv", "json", "xlsx", "xml"):
        if request.method == "GET":
            return templates.TemplateResponse(
                request,
                "error.html",
                {
                    "error_code": "400",
                    "error_message": "Invalid format, format must be json, csv, xml or xlsx",
                    "host_url": host_url,
                },
                status_code=400,
            )
        return JSONResponse(
            {"FormatError": "Invalid format; format must be json, csv, xlsx or xml"},
            status_code=400,
        )

    pipeline = get_stats_pipeline(short_code)
    if validate_emoji_alias(short_code):
        url_data = await EmojiUrlRepository(db["emojis"]).aggregate(pipeline)
    else:
        url_data = await LegacyUrlRepository(db["urls"]).aggregate(pipeline)

    if not url_data:
        if request.method == "GET":
            return templates.TemplateResponse(
                request,
                "error.html",
                {
                    "error_code": "404",
                    "error_message": "URL NOT FOUND",
                    "host_url": host_url,
                },
                status_code=404,
            )
        return JSONResponse(
            {"UrlError": "The requested Url never existed"}, status_code=404
        )

    if url_data.get("password") is not None and password != url_data["password"]:
        if request.method == "POST":
            return JSONResponse({"PasswordError": "Invalid Password"}, status_code=400)
        return templates.TemplateResponse(
            request,
            "error.html",
            {
                "error_code": "400",
                "error_message": "Invalid Password",
                "host_url": host_url,
            },
            status_code=400,
        )

    url_data["short_code"] = short_code
    if url_data.get("max-clicks") is not None:
        url_data["expired"] = url_data["total-clicks"] >= int(url_data["max-clicks"])
    else:
        url_data["expired"] = None

    if url_data.get("expiration-time") is not None:
        expiration_time = convert_to_gmt(url_data["expiration-time"])
        if expiration_time and expiration_time <= datetime.now(timezone.utc):
            url_data["expired"] = True

    (
        url_data["average_daily_clicks"],
        url_data["average_weekly_clicks"],
        url_data["average_monthly_clicks"],
    ) = calculate_click_averages(url_data)

    if url_data["counter"]:
        url_data = add_missing_dates("counter", url_data)
    if url_data.get("unique_counter"):
        url_data = add_missing_dates("unique_counter", url_data)

    if fmt == "json":
        return _export_json(url_data)
    elif fmt == "csv":
        return _export_csv(url_data)
    elif fmt == "xlsx":
        return _export_xlsx(url_data)
    else:
        return _export_xml(url_data)


# ── Export helpers ─────────────────────────────────────────────────────────────


def _export_json(data: dict) -> Response:
    content = json.dumps(data, indent=4, default=str)
    return Response(
        content=content.encode(),
        media_type="application/json",
        headers={"Content-Disposition": 'attachment; filename="spoo-me-export.json"'},
    )


def _export_csv(data: dict) -> Response:
    output = io.BytesIO()

    def write_dict(zipf, filename, dictionary, key_field, value_field):
        with zipf.open(filename, "w") as f:
            with io.TextIOWrapper(f, encoding="utf-8", newline="") as tf:
                writer = csv.writer(tf)
                writer.writerow([key_field, value_field])
                for k, v in dictionary.items():
                    if isinstance(v, dict):
                        v = v.get("counts", 0)
                    writer.writerow([k, v])

    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as zipf:
        general_info = {
            "TOTAL CLICKS": data.get("total-clicks", "N/A"),
            "TOTAL UNIQUE CLICKS": data.get("total_unique_clicks", "N/A"),
            "URL": data.get("url", "N/A"),
            "SHORT CODE": data.get("_id", "N/A"),
            "MAX CLICKS": data.get("max-clicks", "N/A"),
            "EXPIRATION TIME": data.get("expiration-time", "N/A"),
            "PASSWORD": data.get("password", "N/A"),
            "CREATION DATE": data.get("creation-date", "N/A"),
            "CREATION TIME": data.get("creation-time", "N/A"),
            "EXPIRED": data.get("expired", "N/A"),
            "BLOCK BOTS": data.get("block-bots", "N/A"),
            "AVERAGE DAILY CLICKS": data.get("average_daily_clicks", "N/A"),
            "AVERAGE MONTHLY CLICKS": data.get("average_monthly_clicks", "N/A"),
            "AVERAGE WEEKLY CLICKS": data.get("average_weekly_clicks", "N/A"),
            "AVERAGE REDIRECTION TIME (in s)": data.get(
                "average_redirection_time", "N/A"
            ),
            "LAST CLICK": data.get("last-click", "N/A"),
            "LAST CLICK BROWSER": data.get("last-click-browser", "N/A"),
            "LAST CLICK COUNTRY": data.get("last-click-country", "N/A"),
            "LAST CLICK OS": data.get("last-click-os", "N/A"),
        }
        with zipf.open("general_info.csv", "w") as f:
            with io.TextIOWrapper(f, encoding="utf-8", newline="") as tf:
                writer = csv.writer(tf)
                for k, v in general_info.items():
                    writer.writerow([k, v])

        fields = {
            "counter": ("Date", "Count"),
            "browser": ("Browser", "Count"),
            "country": ("Country", "Count"),
            "os_name": ("OS_Name", "Count"),
            "referrer": ("Referrer", "Count"),
            "unique_counter": ("Date", "Count"),
            "unique_browser": ("Browser", "Count"),
            "unique_country": ("Country", "Count"),
            "unique_os_name": ("OS_Name", "Count"),
            "unique_referrer": ("Referrer", "Count"),
            "bots": ("Bot", "Count"),
        }
        for field_name, (kf, vf) in fields.items():
            write_dict(zipf, f"{field_name}.csv", data.get(field_name, {}), kf, vf)

    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/zip",
        headers={
            "Content-Disposition": 'attachment; filename="spoo-me-export-csv.zip"'
        },
    )


def _export_xlsx(data: dict) -> Response:
    output = io.BytesIO()
    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "General_Info"
    general_info = [
        ["TOTAL CLICKS", data.get("total-clicks")],
        ["TOTAL UNIQUE CLICKS", data.get("total_unique_clicks")],
        ["URL", data.get("url")],
        ["SHORT CODE", data.get("_id")],
        ["MAX CLICKS", data.get("max-clicks")],
        ["EXPIRATION TIME", data.get("expiration-time")],
        ["PASSWORD", data.get("password")],
        ["CREATION DATE", data.get("creation-date")],
        ["CREATION TIME", data.get("creation-time")],
        ["EXPIRED", data.get("expired")],
        ["BLOCK BOTS", data.get("block-bots")],
        ["AVERAGE DAILY CLICKS", data.get("average_daily_clicks")],
        ["AVERAGE MONTHLY CLICKS", data.get("average_monthly_clicks")],
        ["AVERAGE WEEKLY CLICKS", data.get("average_weekly_clicks")],
        ["AVERAGE REDIRECTION TIME (in s)", data.get("average_redirection_time")],
        ["LAST CLICK", data.get("last-click")],
        ["LAST CLICK BROWSER", data.get("last-click-browser")],
        ["LAST CLICK OS", data.get("last-click-os")],
        ["LAST CLICK COUNTRY", data.get("last-click-country")],
    ]
    for row in general_info:
        ws.append(row)
    for cell in ws["A"]:
        cell.font = bold
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    for cell in ws["B"]:
        cell.alignment = Alignment(horizontal="right")

    def add_sheet(title, dictionary, col_a, col_b):
        s = wb.create_sheet(title)
        s.append([col_a, col_b])
        s.column_dimensions["A"].width = 20
        for k, v in dictionary.items():
            s.append([k, v])
        for cell in s[1]:
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")

    add_sheet("Browser", data.get("browser", {}), "Browser", "Count")
    add_sheet("Counter", data.get("counter", {}), "Date", "Count")
    add_sheet("Country", data.get("country", {}), "Country", "Count")
    add_sheet("OS_Name", data.get("os_name", {}), "OS_Name", "Count")
    add_sheet("Referrer", data.get("referrer", {}), "Referrer", "Count")
    add_sheet("Unique_Browser", data.get("unique_browser", {}), "Browser", "Count")
    add_sheet("Unique_Counter", data.get("unique_counter", {}), "Date", "Count")
    add_sheet("Unique_Country", data.get("unique_country", {}), "Country", "Count")
    add_sheet("Unique_OS_Name", data.get("unique_os_name", {}), "OS_Name", "Count")
    add_sheet("Unique_Referrer", data.get("unique_referrer", {}), "Referrer", "Count")
    add_sheet("Bots", data.get("bots", {}), "Bot", "Count")

    wb.save(output)
    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="spoo-me-export.xlsx"'},
    )


def _export_xml(data: dict) -> Response:
    xml_bytes = dicttoxml(data)
    return Response(
        content=xml_bytes,
        media_type="application/xml",
        headers={"Content-Disposition": 'attachment; filename="data.xml"'},
    )
