from flask import (
    request,
    render_template,
    redirect,
    url_for,
    jsonify,
    make_response,
    send_file,
)
import re
import string
import random
from datetime import datetime, timedelta, timezone
from pymongo import MongoClient
from dotenv import load_dotenv
import os
from emojies import EMOJIES
from urllib.parse import unquote
import emoji
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import csv
import zipfile
from dicttoxml import dicttoxml
import json
import requests
import validators
import geoip2.errors
import geoip2.database
import pycountry
import functools

load_dotenv(override=True)

MONGO_URI = os.environ["MONGODB_URI"]
CONTACT_WEBHOOK = os.environ["CONTACT_WEBHOOK"]
URL_REPORT_WEBHOOK = os.environ["URL_REPORT_WEBHOOK"]
hcaptcha_secret = os.environ.get("HCAPTCHA_SECRET")

client = MongoClient(MONGO_URI)

try:
    client.admin.command("ping")
    print("\n Pinged your deployment. You successfully connected to MongoDB! \n")
except Exception as e:
    print(e)

db = client["url-shortener"]
collection = db["urls"]
blocked_urls_collection = db["blocked-urls"]
emoji_collection = db["emojis"]
ip_bypasses = db["ip-exceptions"]

with open("bot_user_agents.txt", "r") as file:
    BOT_USER_AGENTS = file.read()
    BOT_USER_AGENTS = [
        i.strip() for i in BOT_USER_AGENTS.split("\n") if i.strip() != ""
    ]


def get_country(ip_address):
    reader = geoip2.database.Reader("misc/GeoLite2-Country.mmdb")
    try:
        response = reader.country(ip_address)
        country = response.country.name
        return country
    except geoip2.errors.AddressNotFoundError:
        return "Unknown"
    finally:
        reader.close()


def get_client_ip():
    if "HTTP_X_FORWARDED_FOR" in request.environ:
        # If the request is proxied, retrieve the IP address from the X-Forwarded-For header
        ip_list = request.environ["HTTP_X_FORWARDED_FOR"].split(",")
        # The client's IP address is typically the first entry in the list
        return ip_list[0].strip()
    else:
        # If the request is not proxied, use the remote address
        return request.environ.get("REMOTE_ADDR", "")


def load_url_by_id(id):
    try:
        url_data = collection.find_one({"_id": id})
    except:
        url_data = None
    return url_data


def add_url_by_id(id, url_data):
    try:
        collection.insert_one({"_id": id, **url_data})
    except:
        pass


def update_url_by_id(id, url_data):
    try:
        collection.update_one({"_id": id}, {"$set": url_data})
    except:
        pass


def check_if_slug_exists(slug):
    try:
        url_data = collection.find_one({"_id": slug})
    except:
        url_data = None
    return url_data is not None


def validate_password(password):
    # Check if the password is at least 8 characters long
    if len(password) < 8:
        return False

    # Check if the password contains a letter, a number, and the allowed special characters
    if not re.search(r"[a-zA-Z]", password):
        return False
    if not re.search(r"\d", password):
        return False
    if not re.search(r"[@.]", password):
        return False

    # Check if there are consecutive special characters
    if re.search(r"[@.]{2}", password):
        return False

    return True


# def validate_url(url):   (OLD DEPRECATED FUNCTION)
#     pattern = re.compile(
#         r"^(https?:\/\/)?(www\.)?[a-zA-Z0-9]+([\-\.]{1}[a-zA-Z0-9]+)*\.[a-zA-Z]{2,6}(\:[0-9]{1,5})?(\/.*)?$"
#     )

#     if "spoo.me" in url:
#         return False

#     if re.fullmatch(pattern, url):
#         return True
#     else:
#         return False


def validate_url(url):
    return (
        validators.url(url, skip_ipv4_addr=True, skip_ipv6_addr=True)
        and "spoo.me" not in url
    )


def validate_blocked_url(url):
    blocked_urls = blocked_urls_collection.find()
    blocked_urls = [doc["_id"] for doc in blocked_urls]

    for blocked_url in blocked_urls:
        if re.search(blocked_url, url):
            return False

    return True


def humanize_number(num):
    magnitude = 0
    while abs(num) >= 1000:
        magnitude += 1
        num /= 1000.0
    return "%d%s+" % (num, ["", "K", "M", "G", "T", "P"][magnitude])


def verify_hcaptcha(token):
    hcaptcha_verify_url = "https://hcaptcha.com/siteverify"

    response = requests.post(
        hcaptcha_verify_url,
        data={
            "response": token,
            "secret": hcaptcha_secret,
        },
    )

    if response.status_code == 200:
        data = response.json()
        return data["success"]
    else:
        return False


def is_positive_integer(value):
    try:
        int(value)
        return True
    except ValueError:
        return False


# custom expiration time is currently really buggy and not ready for production


def validate_expiration_time(expiration_time):
    try:
        expiration_time = datetime.fromisoformat(expiration_time)
        # Check if it's timezone aware
        if expiration_time.tzinfo is None:
            print("timezone not aware")
            return False
        else:
            print("timezone aware")
            print("Expiration Time in GMT: ", expiration_time.astimezone(timezone.utc))
            print(expiration_time.tzinfo)
            # Convert to GMT if it's timezone aware
            expiration_time = expiration_time.astimezone(timezone.utc)
        if expiration_time < datetime.now(timezone.utc) + timedelta(minutes=3):
            print(expiration_time, datetime.now(timezone.utc) + timedelta(minutes=3))
            print("EXPIRATION TIME IN GMT: ", expiration_time)
            print("CURRENT TIME IN GMT: ", datetime.now(timezone.utc))
            print(
                "CURRENT TIME IN GMT + 5: ",
                datetime.now(timezone.utc) + timedelta(minutes=4.5),
            )
            print("less than 5 minutes")
            return False
        return True
    except Exception as e:
        print(e)
        return False


def convert_to_gmt(expiration_time):
    expiration_time = datetime.fromisoformat(expiration_time)
    # Check if it's timezone aware
    if expiration_time.tzinfo is None:
        return None
    else:
        # Convert to GMT if it's timezone aware
        expiration_time = expiration_time.astimezone(timezone.utc)
    return expiration_time


def convert_country_data(data):
    return [{"id": convert_country_name(k), "value": v} for k, v in data.items()]


@functools.lru_cache(maxsize=None)
def convert_country_name(country_name):
    try:
        return pycountry.countries.lookup(country_name).alpha_2
    except LookupError:
        if country_name == "Turkey":
            return "TR"
        elif country_name == "Russia":
            return "RU"
        return "XX"


def generate_short_code():
    letters = string.ascii_lowercase + string.ascii_uppercase + string.digits
    return "".join(random.choice(letters) for i in range(6))


def generate_passkey():
    letters = string.ascii_lowercase + string.ascii_uppercase + string.digits
    return "".join(random.choice(letters) for i in range(22))


def validate_string(string):
    pattern = r"^[a-zA-Z0-9_-]*$"
    return bool(re.search(pattern, string))


def add_missing_dates(key, url_data):
    counter = url_data[key]

    # Get the current date

    # Get the first and last click dates
    first_click_date = url_data["creation-date"]  # next(iter(counter.keys()))
    last_click_date = datetime.now().strftime("%Y-%m-%d")

    # Convert the click dates to datetime objects
    first_date = datetime.strptime(first_click_date, "%Y-%m-%d")
    last_date = datetime.strptime(last_click_date, "%Y-%m-%d")

    # Generate a list of dates between the first and last click dates
    date_range = [
        first_date + timedelta(days=x) for x in range((last_date - first_date).days + 1)
    ]
    all_dates = [date.strftime("%Y-%m-%d") for date in date_range]

    # Add missing dates with a counter value of 0
    for date in all_dates:
        if date not in counter:
            counter[date] = 0

    # Sort the counter dictionary by dates
    sorted_counter = {date: counter[date] for date in sorted(counter.keys())}

    # Update the url_data with the modified counter
    url_data[key] = sorted_counter

    return url_data


def top_four(dictionary):
    if len(dictionary) < 6:
        return dictionary
    sorted_dict = dict(sorted(dictionary.items(), key=lambda x: x[1], reverse=True))
    new_dict = {}
    others = 0
    for i, (key, value) in enumerate(sorted_dict.items()):
        if i < 4:
            new_dict[key] = value
        else:
            others += value

    new_dict["others"] = others
    return new_dict


def calculate_click_averages(data):
    counter = data["counter"]
    total_clicks = data["total-clicks"]
    creation_date = datetime.fromisoformat(data["creation-date"]).date()
    current_date = datetime.now().date()
    link_age = (
        1
        if (current_date - creation_date).days == 0
        else (current_date - creation_date).days
    )

    # Calculate average weekly clicks
    weekly_clicks = sum(counter.values())
    avg_weekly_clicks = round(weekly_clicks / 7, 2)

    # Calculate average daily clicks
    avg_daily_clicks = round(total_clicks / link_age, 2)

    # Calculate average monthly clicks
    avg_monthly_clicks = round(total_clicks / 30, 2)  # Assuming 30 days in a month

    return avg_daily_clicks, avg_weekly_clicks, avg_monthly_clicks


def generate_emoji_alias():
    return "".join(random.choice(EMOJIES) for _ in range(3))


def check_if_emoji_alias_exists(emoji_alias):
    try:
        emoji_data = emoji_collection.find_one({"_id": emoji_alias})
    except:
        emoji_data = None
    return emoji_data is not None


def validate_emoji_alias(alias):
    alias = unquote(alias)
    emoji_list = emoji.emoji_list(alias)
    extracted_emojis = "".join([data["emoji"] for data in emoji_list])
    if len(extracted_emojis) != len(alias) or len(emoji_list) > 15:
        return False
    else:
        return True


def load_emoji_by_alias(alias):
    try:
        emoji_data = emoji_collection.find_one({"_id": alias})
    except:
        emoji_data = None
    return emoji_data


def add_emoji_by_alias(alias, emoji_data):
    try:
        emoji_collection.insert_one({"_id": alias, **emoji_data})
    except:
        pass


def update_emoji_by_alias(alias, emoji_data):
    try:
        emoji_collection.update_one({"_id": alias}, {"$set": emoji_data})
    except:
        pass


def send_report(webhook_uri, short_code, reason, ip_address, host_uri):

    data = {
        "embeds": [
            {
                "title": f"URL Report for `{short_code}`",
                "color": 14177041,
                "url": f"{host_uri}stats/{short_code}",
                "fields": [
                    {"name": "Short Code", "value": f"```{short_code}```"},
                    {"name": "Reason", "value": f"```{reason}```"},
                    {"name": "IP Address", "value": f"```{ip_address}```"},
                ],
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "footer": {
                    "text": "spoo-me",
                    "icon_url": "https://spoo.me/static/images/favicon.png",
                },
            }
        ]
    }

    requests.post(webhook_uri, json=data)


def send_contact_message(webhook_uri, email, message):

    data = {
        "embeds": [
            {
                "title": "New Contact Message ✉️",
                "color": 9103397,
                "fields": [
                    {"name": "Email", "value": f"```{email}```"},
                    {"name": "Message", "value": f"```{message}```"},
                ],
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "footer": {
                    "text": "spoo-me",
                    "icon_url": "https://spoo.me/static/images/favicon.png",
                },
            }
        ]
    }

    requests.post(webhook_uri, json=data)


def export_to_excel(data):
    output = io.BytesIO()
    wb = Workbook()

    # Bold font style
    bold_font = Font(bold=True)

    # General Info Sheet
    ws_general_info = wb.active
    ws_general_info.title = "General_Info"
    general_info = [
        ["TOTAL CLICKS", data["total-clicks"]],
        ["TOTAL UNIQUE CLICKS", data["total_unique_clicks"]],
        ["URL", data["url"]],
        ["SHORT CODE", data["_id"]],
        ["MAX CLICKS", data["max-clicks"]],
        ["EXPIRATION TIME", data["expiration-time"]],
        ["PASSWORD", data["password"]],
        ["CREATION DATE", data["creation-date"]],
        ["EXPIRED", data["expired"]],
        ["BLOCK BOTS", data["block-bots"]],
        ["AVERAGE DAILY CLICKS", data["average_daily_clicks"]],
        ["AVERAGE MONTHLY CLICKS", data["average_monthly_clicks"]],
        ["AVERAGE WEEKLY CLICKS", data["average_weekly_clicks"]],
        ["AVERAGE REDIRECTION TIME (in s)", data["average_redirection_time"]],
        ["LAST CLICK", data["last-click"]],
        ["LAST CLICK BROWSER", data["last-click-browser"]],
        ["LAST CLICK OS", data["last-click-os"]],
    ]
    for row in general_info:
        ws_general_info.append(row)

    # Apply bold font style to the first column (headers)
    for cell in ws_general_info["A"]:
        cell.font = bold_font

    # Set column widths
    ws_general_info.column_dimensions["A"].width = 25
    ws_general_info.column_dimensions["B"].width = 20

    # Align the second column to the right
    for cell in ws_general_info["B"]:
        cell.alignment = Alignment(horizontal="right")

    # Helper function to add data to sheets
    def add_sheet(wb, title, data, columns):
        ws = wb.create_sheet(title)
        ws.append(columns)

        ws.column_dimensions["A"].width = 20

        for key, value in data.items():
            ws.append([key, value])

        # Apply bold font style to the first row (headers) and center align them
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")

    # Adding other sheets
    add_sheet(wb, "Browser", data["browser"], ["Browser", "Count"])
    add_sheet(wb, "Counter", data["counter"], ["Date", "Count"])
    add_sheet(wb, "Country", data["country"], ["Country", "Count"])
    add_sheet(wb, "OS_Name", data["os_name"], ["OS_Name", "Count"])
    add_sheet(wb, "Referrer", data["referrer"], ["Referrer", "Count"])
    add_sheet(wb, "Unique_Browser", data["unique_browser"], ["Browser", "Count"])
    add_sheet(wb, "Unique_Counter", data["unique_counter"], ["Date", "Count"])
    add_sheet(wb, "Unique_Country", data["unique_country"], ["Country", "Count"])
    add_sheet(wb, "Unique_OS_Name", data["unique_os_name"], ["OS_Name", "Count"])
    add_sheet(wb, "Unique_Referrer", data["unique_referrer"], ["Referrer", "Count"])
    add_sheet(wb, "Bots", data["bots"], ["Bot", "Count"])

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="spoo-me-export.xlsx",
    )


def export_to_csv(data):
    output = io.BytesIO()

    # Helper function to write a dictionary to a CSV file in the zip
    def write_dict_to_csv(zipf, filename, dictionary, key_field, value_field):
        with zipf.open(filename, "w") as file:
            # Use TextIOWrapper to handle encoding and newline characters properly
            with io.TextIOWrapper(file, encoding="utf-8", newline="") as text_file:
                writer = csv.writer(text_file)
                writer.writerow([key_field, value_field])
                for key, value in dictionary.items():
                    if isinstance(value, dict):
                        # Handle nested dictionaries
                        value = value.get("counts", 0)
                    writer.writerow([key, value])

    # Create a zip file in memory
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as zipf:
        # Write general info CSV
        general_info = {
            "TOTAL CLICKS": data.get("total-clicks", "N/A"),
            "TOTAL UNIQUE CLICKS": data.get("total_unique_clicks", "N/A"),
            "URL": data.get("url", "N/A"),
            "SHORT CODE": data.get("_id", "N/A"),
            "MAX CLICKS": data.get("max-clicks", "N/A"),
            "EXPIRATION TIME": data.get("expiration-time", "N/A"),
            "PASSWORD": data.get("password", "N/A"),
            "CREATION DATE": data.get("creation-date", "N/A"),
            "CREATION TIME": data.get("creation-time", "N/A"),
            "CREATION IP ADDRESS": data.get("creation-ip-address", "N/A"),
            "EXPIRED": data.get("expired", "N/A"),
            "BLOCK BOTS": data.get("block-bots", "N/A"),
            "AVERAGE DAILY CLICKS": data.get("average_daily_clicks", "N/A"),
            "AVERAGE MONTHLY CLICKS": data.get("average_monthly_clicks", "N/A"),
            "AVERAGE WEEKLY CLICKS": data.get("average_weekly_clicks", "N/A"),
            "AVERAGE REDIRECTION TIME (in s)": data.get(
                "average_redirection_time", "N/A"
            ),
            "LAST CLICK": data.get("last-click", "N/A"),
            "LAST CLICK BROWSER": data.get("last-click-browser", "N/A"),
            "LAST CLICK COUNTRY": data.get("last-click-country", "N/A"),
            "LAST CLICK OS": data.get("last-click-os", "N/A"),
        }
        with zipf.open("general_info.csv", "w") as file:
            with io.TextIOWrapper(file, encoding="utf-8", newline="") as text_file:
                writer = csv.writer(text_file)
                for key, value in general_info.items():
                    writer.writerow([key, value])

        # Write other CSV files dynamically
        fields = {
            "counter": ("Date", "Count"),
            "browser": ("Browser", "Count"),
            "country": ("Country", "Count"),
            "os_name": ("OS_Name", "Count"),
            "referrer": ("Referrer", "Count"),
            "unique_counter": ("Date", "Count"),
            "unique_browser": ("Browser", "Count"),
            "unique_country": ("Country", "Count"),
            "unique_os_name": ("OS_Name", "Count"),
            "unique_referrer": ("Referrer", "Count"),
            "bots": ("Bot", "Count"),
        }

        for field_name, (key_field, value_field) in fields.items():
            write_dict_to_csv(
                zipf,
                f"{field_name}.csv",
                data.get(field_name, {}),
                key_field,
                value_field,
            )

    output.seek(0)

    return send_file(
        output,
        mimetype="application/zip",
        as_attachment=True,
        download_name="spoo-me-export-csv.zip",
    )


def export_to_json(data):
    output = io.StringIO()

    json.dump(data, output, indent=4)
    output.seek(0)

    output_bytes = io.BytesIO(output.getvalue().encode())

    return send_file(
        output_bytes,
        mimetype="application/json",
        as_attachment=True,
        download_name="spoo-me-export.json",
    )


def export_to_xml(data):
    # Convert dictionary to XML
    xml = dicttoxml(data)

    # Create BytesIO object and write XML data to it
    output = io.BytesIO()
    output.write(xml)
    output.seek(0)

    # Send file
    return send_file(
        output, mimetype="application/xml", as_attachment=True, download_name="data.xml"
    )
