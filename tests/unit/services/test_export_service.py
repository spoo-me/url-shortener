"""Unit tests for Phase 9 — ExportService and export formatters."""

from __future__ import annotations

import json
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from errors import NotFoundError, ValidationError
from schemas.dto.requests.stats import ExportQuery
from schemas.results import ExportResult
from services.export import ExportService, default_formatters
from services.export.formatters import (
    CsvFormatter,
    JsonFormatter,
    XlsxFormatter,
    XmlFormatter,
)
from services.stats_service import StatsService

# ── Constants ─────────────────────────────────────────────────────────────────

NOW = datetime(2024, 6, 15, 12, 0, 0, tzinfo=timezone.utc)
START = datetime(2024, 6, 8, 12, 0, 0, tzinfo=timezone.utc)

NOW_ISO = NOW.isoformat()
START_ISO = START.isoformat()

SAMPLE_STATS = {
    "scope": "anon",
    "short_code": "abc",
    "filters": {},
    "group_by": ["browser"],
    "timezone": "UTC",
    "metrics": {
        "clicks_by_browser": [
            {"browser": "Chrome", "clicks": 80},
            {"browser": "Firefox", "clicks": 20},
        ]
    },
    "summary": {
        "total_clicks": 100,
        "unique_clicks": 40,
        "first_click": "2024-06-08T12:00:00+00:00",
        "last_click": "2024-06-15T12:00:00+00:00",
        "avg_redirection_time": 120.5,
    },
    "time_range": {
        "start_date": "2024-06-08T12:00:00+00:00",
        "end_date": "2024-06-15T12:00:00+00:00",
    },
    "generated_at": "2024-06-15T12:00:00+00:00",
    "api_version": "v1",
}


def _export_query(fmt="json"):
    """Build an ExportQuery with sensible defaults for tests."""
    return ExportQuery(
        format=fmt,
        scope="anon",
        short_code="abc",
        start_date=START_ISO,
        end_date=NOW_ISO,
        group_by="browser",
        metrics="clicks",
        timezone="UTC",
    )


# ── Helpers ───────────────────────────────────────────────────────────────────


def make_service(stats_data=None):
    stats_svc = AsyncMock(spec=StatsService)
    stats_svc.query = AsyncMock(return_value=stats_data or SAMPLE_STATS)
    return ExportService(
        stats_service=stats_svc, formatters=default_formatters()
    ), stats_svc


# ── Tests: individual formatters ──────────────────────────────────────────────


class TestFormatters:
    def test_json_formatter_mimetype_and_filename(self):
        f = JsonFormatter()
        assert f.mimetype == "application/json"
        assert f.filename == "spoo-me-export.json"

    def test_json_formatter_produces_valid_json(self):
        content = JsonFormatter().serialize(SAMPLE_STATS)
        parsed = json.loads(content)
        assert parsed["summary"]["total_clicks"] == 100

    def test_xml_formatter_mimetype_and_filename(self):
        f = XmlFormatter()
        assert f.mimetype == "application/xml"
        assert f.filename == "spoo-me-export.xml"

    def test_xml_formatter_produces_bytes(self):
        content = XmlFormatter().serialize(SAMPLE_STATS)
        assert isinstance(content, bytes) and len(content) > 0

    def test_csv_formatter_mimetype_and_filename(self):
        f = CsvFormatter()
        assert f.mimetype == "application/zip"
        assert f.filename == "spoo-me-export-csv.zip"

    def test_csv_formatter_produces_valid_zip(self):
        content = CsvFormatter().serialize(SAMPLE_STATS)
        assert zipfile.is_zipfile(BytesIO(content))

    def test_csv_formatter_zip_contains_summary(self):
        content = CsvFormatter().serialize(SAMPLE_STATS)
        with zipfile.ZipFile(BytesIO(content)) as zf:
            assert "summary.csv" in zf.namelist()

    def test_csv_formatter_zip_contains_dimension_file(self):
        content = CsvFormatter().serialize(SAMPLE_STATS)
        with zipfile.ZipFile(BytesIO(content)) as zf:
            assert "clicks_by_browser.csv" in zf.namelist()

    def test_csv_formatter_summary_contains_total_clicks(self):
        content = CsvFormatter().serialize(SAMPLE_STATS)
        with zipfile.ZipFile(BytesIO(content)) as zf, zf.open("summary.csv") as f:
            csv_text = f.read().decode("utf-8")
        assert "total_clicks" in csv_text
        assert "100" in csv_text

    def test_csv_formatter_empty_metrics_produces_summary_only(self):
        data = {**SAMPLE_STATS, "metrics": {}}
        content = CsvFormatter().serialize(data)
        with zipfile.ZipFile(BytesIO(content)) as zf:
            assert zf.namelist() == ["summary.csv"]

    def test_xlsx_formatter_mimetype_and_filename(self):
        f = XlsxFormatter()
        assert "spreadsheetml" in f.mimetype
        assert f.filename == "spoo-me-export.xlsx"

    def test_xlsx_formatter_produces_valid_workbook(self):
        content = XlsxFormatter().serialize(SAMPLE_STATS)
        wb = load_workbook(BytesIO(content))
        assert "Summary" in wb.sheetnames

    def test_xlsx_formatter_has_dimension_sheet(self):
        content = XlsxFormatter().serialize(SAMPLE_STATS)
        wb = load_workbook(BytesIO(content))
        assert "clicks_by_browser" in wb.sheetnames

    def test_xlsx_formatter_summary_contains_total_clicks_row(self):
        content = XlsxFormatter().serialize(SAMPLE_STATS)
        wb = load_workbook(BytesIO(content))
        ws = wb["Summary"]
        cell_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "total_clicks" in cell_values


# ── Tests: ExportService — format validation ──────────────────────────────────


class TestFormatValidation:
    @pytest.mark.asyncio
    async def test_unknown_format_raises_validation_error(self):
        svc, _ = make_service()
        # ExportQuery validates format, so we need to bypass pydantic validation
        # by constructing a query with a valid format, then patching it
        query = _export_query("json")
        # Manually override the format to simulate an invalid one getting through
        object.__setattr__(query, "format", "pdf")
        with pytest.raises(ValidationError, match="invalid format"):
            await svc.export(query=query, owner_id=None)

    @pytest.mark.asyncio
    async def test_known_formats_do_not_raise(self):
        for fmt in ("json", "xml", "csv", "xlsx"):
            svc, _ = make_service()
            result = await svc.export(query=_export_query(fmt), owner_id=None)
            assert len(result.content) > 0, f"{fmt} produced empty content"


# ── Tests: ExportService — return type ───────────────────────────────────────


class TestReturnType:
    @pytest.mark.asyncio
    async def test_returns_export_result(self):
        svc, _ = make_service()
        result = await svc.export(query=_export_query("json"), owner_id=None)
        assert isinstance(result, ExportResult)

    @pytest.mark.asyncio
    async def test_mimetype_and_filename_come_from_formatter(self):
        svc, _ = make_service()
        result = await svc.export(query=_export_query("json"), owner_id=None)
        assert result.mimetype == "application/json"
        assert result.filename == "spoo-me-export.json"


# ── Tests: ExportService — delegation ────────────────────────────────────────


class TestDelegation:
    @pytest.mark.asyncio
    async def test_calls_stats_service_query_once(self):
        svc, stats_svc = make_service()
        await svc.export(query=_export_query("json"), owner_id=None)
        stats_svc.query.assert_awaited_once()

    @pytest.mark.asyncio
    async def test_passes_query_to_stats_service(self):
        svc, stats_svc = make_service()
        query = _export_query("json")
        await svc.export(query=query, owner_id=None)
        stats_svc.query.assert_awaited_once()
        call_args = stats_svc.query.call_args
        assert call_args[0][0] is query, "must forward the exact query object"
        assert call_args[0][1] is None, "must forward owner_id as-is"

    @pytest.mark.asyncio
    async def test_stats_service_error_propagates(self):
        svc, stats_svc = make_service()
        stats_svc.query.side_effect = NotFoundError("not found")
        with pytest.raises(NotFoundError):
            await svc.export(query=_export_query("json"), owner_id=None)

    @pytest.mark.asyncio
    async def test_custom_formatter_is_called(self):
        """ExportService dispatches to whatever formatter is in the registry."""
        custom_fmt = MagicMock()
        custom_fmt.serialize.return_value = b"custom"
        custom_fmt.mimetype = "application/custom"
        custom_fmt.filename = "out.custom"

        stats_svc = AsyncMock(spec=StatsService)
        stats_svc.query = AsyncMock(return_value=SAMPLE_STATS)

        svc = ExportService(
            stats_service=stats_svc,
            formatters={"custom": custom_fmt},
        )
        # Build a query with a valid format then override to "custom"
        query = _export_query("json")
        object.__setattr__(query, "format", "custom")
        result = await svc.export(query=query, owner_id=None)
        custom_fmt.serialize.assert_called_once_with(SAMPLE_STATS)
        assert result.content == b"custom"
        assert result.mimetype == "application/custom"
        assert result.filename == "out.custom"
