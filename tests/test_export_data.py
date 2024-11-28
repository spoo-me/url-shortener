import io
from openpyxl import load_workbook
from utils.export_utils import (
    export_to_excel,
    export_to_csv,
    export_to_json,
    export_to_xml,
)
import csv
import json
import zipfile
from dicttoxml import dicttoxml

# Mock data
data = {
    "_id": "abc123",
    "total-clicks": 100,
    "total_unique_clicks": 80,
    "url": "https://example.com",
    "short-code": "abc123",
    "max-clicks": 10,
    "expiration-time": "2024-12-31",
    "password": "password123",
    "creation-date": "2024-01-01",
    "creation-time": "12:00:00",
    "expired": False,
    "block-bots": True,
    "average_daily_clicks": 5,
    "average_monthly_clicks": 150,
    "average_weekly_clicks": 35,
    "average_redirection_time": 0.5,
    "last-click": "2024-10-15",
    "last-click-browser": "Chrome",
    "last-click-os": "Windows",
    "last-click-country": "US",
    "browser": {"Chrome": 50, "Firefox": 30},
    "counter": {"2024-10-01": 10, "2024-10-02": 20},
    "country": {"US": 60, "DE": 20},
    "os_name": {"Windows": 70, "MacOS": 10},
    "referrer": {"google.com": 40, "bing.com": 10},
    "unique_browser": {"Chrome": 40, "Firefox": 20},
    "unique_counter": {"2024-10-01": 8, "2024-10-02": 12},
    "unique_country": {"US": 50, "DE": 15},
    "unique_os_name": {"Windows": 60, "MacOS": 5},
    "unique_referrer": {"google.com": 30, "bing.com": 5},
    "bots": {"Googlebot": 5, "Bingbot": 2},
}


def test_export_to_excel(mocker):
    # Mock send_file
    mock_send_file = mocker.patch("utils.export_utils.send_file")

    # Call the function
    export_to_excel(data)

    # Assert send_file was called
    assert mock_send_file.called
    assert mock_send_file.call_count == 1

    # Get the file-like object passed to send_file
    file_obj = mock_send_file.call_args[0][0]

    # Load the workbook from the file-like object
    wb = load_workbook(file_obj)

    # Check the sheet names
    expected_sheets = [
        "General_Info",
        "Browser",
        "Counter",
        "Country",
        "OS_Name",
        "Referrer",
        "Unique_Browser",
        "Unique_Counter",
        "Unique_Country",
        "Unique_OS_Name",
        "Unique_Referrer",
        "Bots",
    ]
    assert wb.sheetnames == expected_sheets

    # Check the content of the General_Info sheet
    ws_general_info = wb["General_Info"]
    general_info = [
        ["TOTAL CLICKS", data["total-clicks"]],
        ["TOTAL UNIQUE CLICKS", data["total_unique_clicks"]],
        ["URL", data["url"]],
        ["SHORT CODE", data["short-code"]],
        ["MAX CLICKS", data["max-clicks"]],
        ["EXPIRATION TIME", data["expiration-time"]],
        ["PASSWORD", data["password"]],
        ["CREATION DATE", data["creation-date"]],
        ["CREATION TIME", data["creation-time"]],
        ["EXPIRED", data["expired"]],
        ["BLOCK BOTS", data["block-bots"]],
        ["AVERAGE DAILY CLICKS", data["average_daily_clicks"]],
        ["AVERAGE MONTHLY CLICKS", data["average_monthly_clicks"]],
        ["AVERAGE WEEKLY CLICKS", data["average_weekly_clicks"]],
        ["AVERAGE REDIRECTION TIME (in s)", data["average_redirection_time"]],
        ["LAST CLICK", data["last-click"]],
        ["LAST CLICK BROWSER", data["last-click-browser"]],
        ["LAST CLICK OS", data["last-click-os"]],
        ["LAST CLICK COUNTRY", data["last-click-country"]],
    ]
    for row_idx, row in enumerate(general_info, start=1):
        assert ws_general_info.cell(row=row_idx, column=1).value == row[0]
        assert ws_general_info.cell(row=row_idx, column=2).value == row[1]

    # Check the content of other sheets
    def check_sheet_content(sheet_name, expected_data):
        ws = wb[sheet_name]
        for row_idx, (key, value) in enumerate(expected_data.items(), start=2):
            assert ws.cell(row=row_idx, column=1).value == key
            assert ws.cell(row=row_idx, column=2).value == value

    check_sheet_content("Browser", data["browser"])
    check_sheet_content("Counter", data["counter"])
    check_sheet_content("Country", data["country"])
    check_sheet_content("OS_Name", data["os_name"])
    check_sheet_content("Referrer", data["referrer"])
    check_sheet_content("Unique_Browser", data["unique_browser"])
    check_sheet_content("Unique_Counter", data["unique_counter"])
    check_sheet_content("Unique_Country", data["unique_country"])
    check_sheet_content("Unique_OS_Name", data["unique_os_name"])
    check_sheet_content("Unique_Referrer", data["unique_referrer"])
    check_sheet_content("Bots", data["bots"])


def test_export_to_csv(mocker):
    # Mock send_file
    mock_send_file = mocker.patch("utils.export_utils.send_file")

    # Call the function
    export_to_csv(data)

    # Assert send_file was called
    assert mock_send_file.called
    assert mock_send_file.call_count == 1

    # Get the file-like object passed to send_file
    file_obj = mock_send_file.call_args[0][0]

    # Load the zip file from the file-like object
    with zipfile.ZipFile(file_obj, "r") as zipf:
        # Check the list of files in the zip
        expected_files = [
            "general_info.csv",
            "counter.csv",
            "browser.csv",
            "country.csv",
            "os_name.csv",
            "referrer.csv",
            "unique_counter.csv",
            "unique_browser.csv",
            "unique_country.csv",
            "unique_os_name.csv",
            "unique_referrer.csv",
            "bots.csv",
        ]
        assert sorted(zipf.namelist()) == sorted(expected_files)

        # Check the content of the general_info.csv file
        with zipf.open("general_info.csv") as file:
            with io.TextIOWrapper(file, encoding="utf-8") as text_file:
                reader = csv.reader(text_file)
                general_info = {
                    "TOTAL CLICKS": str(data["total-clicks"]),
                    "TOTAL UNIQUE CLICKS": str(data["total_unique_clicks"]),
                    "URL": data["url"],
                    "SHORT CODE": data["_id"],
                    "MAX CLICKS": str(data["max-clicks"]),
                    "EXPIRATION TIME": data["expiration-time"],
                    "PASSWORD": data["password"],
                    "CREATION DATE": data["creation-date"],
                    "CREATION TIME": data["creation-time"],
                    "EXPIRED": str(data["expired"]),
                    "BLOCK BOTS": str(data["block-bots"]),
                    "AVERAGE DAILY CLICKS": str(data["average_daily_clicks"]),
                    "AVERAGE MONTHLY CLICKS": str(data["average_monthly_clicks"]),
                    "AVERAGE WEEKLY CLICKS": str(data["average_weekly_clicks"]),
                    "AVERAGE REDIRECTION TIME (in s)": str(
                        data["average_redirection_time"]
                    ),
                    "LAST CLICK": data["last-click"],
                    "LAST CLICK BROWSER": data["last-click-browser"],
                    "LAST CLICK COUNTRY": data["last-click-country"],
                    "LAST CLICK OS": data["last-click-os"],
                }
                for row in reader:
                    key, value = row
                    assert general_info[key] == value

        # Check the content of other CSV files
        def check_csv_content(filename, expected_data):
            with zipf.open(filename) as file:
                with io.TextIOWrapper(file, encoding="utf-8") as text_file:
                    reader = csv.reader(text_file)
                    for row in reader:
                        key, value = row
                        assert expected_data[key] == int(value)

        check_csv_content("counter.csv", data["counter"])
        check_csv_content("browser.csv", data["browser"])
        check_csv_content("country.csv", data["country"])
        check_csv_content("os_name.csv", data["os_name"])
        check_csv_content("referrer.csv", data["referrer"])
        check_csv_content("unique_counter.csv", data["unique_counter"])
        check_csv_content("unique_browser.csv", data["unique_browser"])
        check_csv_content("unique_country.csv", data["unique_country"])
        check_csv_content("unique_os_name.csv", data["unique_os_name"])
        check_csv_content("unique_referrer.csv", data["unique_referrer"])
        check_csv_content("bots.csv", data["bots"])


def test_export_to_json(mocker):
    # Mock send_file
    mock_send_file = mocker.patch("utils.export_utils.send_file")

    # Call the function
    export_to_json(data)

    # Assert send_file was called
    assert mock_send_file.called
    assert mock_send_file.call_count == 1

    # Get the file-like object passed to send_file
    file_obj = mock_send_file.call_args[0][0]

    # Read the JSON data from the file-like object
    json_data = json.load(file_obj)

    # Check if the JSON data matches the input data
    assert json_data == data


def test_export_to_xml(mocker):
    # Mock send_file
    mock_send_file = mocker.patch("utils.export_utils.send_file")

    # Call the function
    export_to_xml(data)

    # Assert send_file was called
    assert mock_send_file.called
    assert mock_send_file.call_count == 1

    # Get the file-like object passed to send_file
    file_obj = mock_send_file.call_args[0][0]

    # Read the XML data from the file-like object
    xml_data = file_obj.read()

    # Convert the input data to XML
    expected_xml = dicttoxml(data)

    # Check if the XML data matches the expected XML
    assert xml_data == expected_xml
