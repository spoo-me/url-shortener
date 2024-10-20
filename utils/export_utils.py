import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import csv
import zipfile
from dicttoxml import dicttoxml
import json
from flask import send_file


def export_to_excel(data):
    output = io.BytesIO()
    wb = Workbook()

    # Bold font style
    bold_font = Font(bold=True)

    # General Info Sheet
    ws_general_info = wb.active
    ws_general_info.title = "General_Info"
    general_info = [
        ["TOTAL CLICKS", data["total-clicks"]],
        ["TOTAL UNIQUE CLICKS", data["total_unique_clicks"]],
        ["URL", data["url"]],
        ["SHORT CODE", data["_id"]],
        ["MAX CLICKS", data["max-clicks"]],
        ["EXPIRATION TIME", data["expiration-time"]],
        ["PASSWORD", data["password"]],
        ["CREATION DATE", data["creation-date"]],
        ["CREATION TIME", data["creation-time"]],
        ["EXPIRED", data["expired"]],
        ["BLOCK BOTS", data["block-bots"]],
        ["AVERAGE DAILY CLICKS", data["average_daily_clicks"]],
        ["AVERAGE MONTHLY CLICKS", data["average_monthly_clicks"]],
        ["AVERAGE WEEKLY CLICKS", data["average_weekly_clicks"]],
        ["AVERAGE REDIRECTION TIME (in s)", data["average_redirection_time"]],
        ["LAST CLICK", data["last-click"]],
        ["LAST CLICK BROWSER", data["last-click-browser"]],
        ["LAST CLICK OS", data["last-click-os"]],
        ["LAST CLICK COUNTRY", data["last-click-country"]],
    ]
    for row in general_info:
        ws_general_info.append(row)

    # Apply bold font style to the first column (headers)
    for cell in ws_general_info["A"]:
        cell.font = bold_font

    # Set column widths
    ws_general_info.column_dimensions["A"].width = 25
    ws_general_info.column_dimensions["B"].width = 20

    # Align the second column to the right
    for cell in ws_general_info["B"]:
        cell.alignment = Alignment(horizontal="right")

    # Helper function to add data to sheets
    def add_sheet(wb, title, data, columns):
        ws = wb.create_sheet(title)
        ws.append(columns)

        ws.column_dimensions["A"].width = 20

        for key, value in data.items():
            ws.append([key, value])

        # Apply bold font style to the first row (headers) and center align them
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")

    # Adding other sheets
    add_sheet(wb, "Browser", data["browser"], ["Browser", "Count"])
    add_sheet(wb, "Counter", data["counter"], ["Date", "Count"])
    add_sheet(wb, "Country", data["country"], ["Country", "Count"])
    add_sheet(wb, "OS_Name", data["os_name"], ["OS_Name", "Count"])
    add_sheet(wb, "Referrer", data["referrer"], ["Referrer", "Count"])
    add_sheet(wb, "Unique_Browser", data["unique_browser"], ["Browser", "Count"])
    add_sheet(wb, "Unique_Counter", data["unique_counter"], ["Date", "Count"])
    add_sheet(wb, "Unique_Country", data["unique_country"], ["Country", "Count"])
    add_sheet(wb, "Unique_OS_Name", data["unique_os_name"], ["OS_Name", "Count"])
    add_sheet(wb, "Unique_Referrer", data["unique_referrer"], ["Referrer", "Count"])
    add_sheet(wb, "Bots", data["bots"], ["Bot", "Count"])

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="spoo-me-export.xlsx",
    )


def export_to_csv(data):
    output = io.BytesIO()

    # Helper function to write a dictionary to a CSV file in the zip
    def write_dict_to_csv(zipf, filename, dictionary, key_field, value_field):
        with zipf.open(filename, "w") as file:
            # Use TextIOWrapper to handle encoding and newline characters properly
            with io.TextIOWrapper(file, encoding="utf-8", newline="") as text_file:
                writer = csv.writer(text_file)
                writer.writerow([key_field, value_field])
                for key, value in dictionary.items():
                    if isinstance(value, dict):
                        # Handle nested dictionaries
                        value = value.get("counts", 0)
                    writer.writerow([key, value])

    # Create a zip file in memory
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as zipf:
        # Write general info CSV
        general_info = {
            "TOTAL CLICKS": data.get("total-clicks", "N/A"),
            "TOTAL UNIQUE CLICKS": data.get("total_unique_clicks", "N/A"),
            "URL": data.get("url", "N/A"),
            "SHORT CODE": data.get("_id", "N/A"),
            "MAX CLICKS": data.get("max-clicks", "N/A"),
            "EXPIRATION TIME": data.get("expiration-time", "N/A"),
            "PASSWORD": data.get("password", "N/A"),
            "CREATION DATE": data.get("creation-date", "N/A"),
            "CREATION TIME": data.get("creation-time", "N/A"),
            "EXPIRED": data.get("expired", "N/A"),
            "BLOCK BOTS": data.get("block-bots", "N/A"),
            "AVERAGE DAILY CLICKS": data.get("average_daily_clicks", "N/A"),
            "AVERAGE MONTHLY CLICKS": data.get("average_monthly_clicks", "N/A"),
            "AVERAGE WEEKLY CLICKS": data.get("average_weekly_clicks", "N/A"),
            "AVERAGE REDIRECTION TIME (in s)": data.get(
                "average_redirection_time", "N/A"
            ),
            "LAST CLICK": data.get("last-click", "N/A"),
            "LAST CLICK BROWSER": data.get("last-click-browser", "N/A"),
            "LAST CLICK COUNTRY": data.get("last-click-country", "N/A"),
            "LAST CLICK OS": data.get("last-click-os", "N/A"),
        }
        with zipf.open("general_info.csv", "w") as file:
            with io.TextIOWrapper(file, encoding="utf-8", newline="") as text_file:
                writer = csv.writer(text_file)
                for key, value in general_info.items():
                    writer.writerow([key, value])

        # Write other CSV files dynamically
        fields = {
            "counter": ("Date", "Count"),
            "browser": ("Browser", "Count"),
            "country": ("Country", "Count"),
            "os_name": ("OS_Name", "Count"),
            "referrer": ("Referrer", "Count"),
            "unique_counter": ("Date", "Count"),
            "unique_browser": ("Browser", "Count"),
            "unique_country": ("Country", "Count"),
            "unique_os_name": ("OS_Name", "Count"),
            "unique_referrer": ("Referrer", "Count"),
            "bots": ("Bot", "Count"),
        }

        for field_name, (key_field, value_field) in fields.items():
            write_dict_to_csv(
                zipf,
                f"{field_name}.csv",
                data.get(field_name, {}),
                key_field,
                value_field,
            )

    output.seek(0)

    return send_file(
        output,
        mimetype="application/zip",
        as_attachment=True,
        download_name="spoo-me-export-csv.zip",
    )


def export_to_json(data):
    output = io.StringIO()

    json.dump(data, output, indent=4)
    output.seek(0)

    output_bytes = io.BytesIO(output.getvalue().encode())

    return send_file(
        output_bytes,
        mimetype="application/json",
        as_attachment=True,
        download_name="spoo-me-export.json",
    )


def export_to_xml(data):
    # Convert dictionary to XML
    xml = dicttoxml(data)

    # Create BytesIO object and write XML data to it
    output = io.BytesIO()
    output.write(xml)
    output.seek(0)

    # Send file
    return send_file(
        output, mimetype="application/xml", as_attachment=True, download_name="data.xml"
    )
